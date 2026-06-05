"""Agent-team ticket test harness (ad-hoc, Phase-4 live round-trip).

Pipeline per ticket:
  1. Read CS-agent-entered properties from the category CSV.
  2. Fetch from Freshdesk PRODUCTION (read-only): customer first message
     (ticket description_text) + first public CS-agent reply.
  3. Run the REAL cs-agent-team (cs-lead + subagents + .claude/ hooks) via the
     `claude` CLI headless (DRY_RUN — nothing posted to Freshdesk), asking it to
     report both its classification `properties` and its final `verdict`.
  4. Persist one JSON record per ticket to a gitignored data file.

Then `xlsx` builds test-tickets.xlsx (one sheet per ticket).

PII discipline: customer message / CS reply / AI reply are written ONLY to the
gitignored data file + xlsx (both gitignored). Never printed raw to stdout.

Usage:
  .venv/bin/python scripts/test_tickets_run.py collect --per-cat 10
  .venv/bin/python scripts/test_tickets_run.py collect --ticket 7731117 --category inquiry   # smoke
  .venv/bin/python scripts/test_tickets_run.py xlsx
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import os
import sys
import time
from pathlib import Path
from typing import Any

import httpx

_REPO_ROOT = Path(__file__).parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

# Reuse the production runner machinery (redaction, pre-screen, CLI invoke, parse).
# D-32: _post_screen_draft, _sanitize_ticket_id, _state_file_path removed (deleted guard hooks gone).
# D-33: verdict is always action=draft; no escalate=no-draft path.
from scripts.cs_team_demo import (  # noqa: E402
    _CLAUDE_CLI,
    _parse_verdict,
    _pre_screen_ticket,
    redact_text,
    settings,
)

import uuid  # noqa: E402 — still used for unique run IDs in run_ai_team

_TICKET_DIR = _REPO_ROOT / ".planning" / "phases" / \
    "01-knowledge-survey-conflict-inventory" / "snapshots" / "confluence" / "ticket-sample"
_CSV = {
    "change_request": _TICKET_DIR / "change-request_tickets.csv",
    "complaint": _TICKET_DIR / "complaint_tickets.csv",
    "inquiry": _TICKET_DIR / "inquiry_tickets.csv",
}
_DATA_PATH = _REPO_ROOT / ".test-tickets-data.jsonl"
_XLSX_PATH = _REPO_ROOT / "test-tickets.xlsx"

# CS property columns to surface as rows (semantic-priority first); any other
# non-empty column is appended after these in CSV order.
_PRIORITY_PROPS = [
    "Subject", "Status", "Priority", "Source", "Type", "Group", "Tags",
    "Order", "Related orders", "Summary", "Level_in", "Customer_Request",
    "Product_line", "Customer_Feedback", "Feedback_Issue", "Additional_Feedback",
    "Additional_Feedback_Issue", "Rootcause", "Rootcause_type", "Section_Flow",
    "Product_label", "Escalation level", "Flow", "STEP", "Handler",
    "Package_status", "Resolution status",
]


# ---------------------------------------------------------------------------
# Freshdesk (read-only) — load .env.prd directly (override stale shell env)
# ---------------------------------------------------------------------------

def _load_env_prd() -> dict[str, str]:
    # Walk up from _REPO_ROOT to find .env.prd — handles both main repo and worktree layouts
    # (worktrees share credentials with the main repo; the file lives in the main repo root).
    candidate = _REPO_ROOT / ".env.prd"
    if not candidate.exists():
        for parent in _REPO_ROOT.parents:
            alt = parent / ".env.prd"
            if alt.exists():
                candidate = alt
                break
    if not candidate.exists():
        raise FileNotFoundError(
            f".env.prd not found in {_REPO_ROOT} or any parent directory. "
            "Copy it from the main repo root or set FRESHDESK_DOMAIN / "
            "FRESHDESK_API_KEY in your environment."
        )
    env: dict[str, str] = {}
    for line in candidate.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def fetch_conversation(client: httpx.Client, domain: str, key: str, tid: str) -> dict:
    """Return {customer_msg, cs_reply, error}. No raw PII printed by caller."""
    out: dict[str, Any] = {"customer_msg": "", "cs_reply": "", "error": ""}
    try:
        t = client.get(f"https://{domain}/api/v2/tickets/{tid}",
                       auth=(key, "X"), timeout=30)
        if t.status_code != 200:
            out["error"] = f"ticket HTTP {t.status_code}"
            return out
        tj = t.json()
        out["customer_msg"] = (tj.get("description_text") or "").strip()[:6000]

        conv = client.get(f"https://{domain}/api/v2/tickets/{tid}/conversations",
                          auth=(key, "X"), timeout=30)
        if conv.status_code == 200:
            pub = [c for c in conv.json()
                   if c.get("incoming") is False and c.get("private") is False]
            pub.sort(key=lambda c: c.get("created_at", ""))
            if pub:
                out["cs_reply"] = (pub[0].get("body_text") or "").strip()[:6000]
    except Exception as exc:  # noqa: BLE001
        out["error"] = str(exc)
    return out


# ---------------------------------------------------------------------------
# AI agent team — real cs-lead via claude CLI, combined properties+verdict prompt
# ---------------------------------------------------------------------------

def _build_test_prompt(ticket: dict, selless: dict | None = None) -> str:
    body = redact_text(ticket.get("body", ""))
    subject = redact_text(ticket.get("subject", ""))
    order_ref = redact_text(ticket.get("order_ref", ""))
    # D-34: real Selless order data grounds the reply (whitelisted fields only).
    # When present, the AI MUST use it (status/tracking/variant) and must NOT ask
    # the customer for the order number or fabricate order facts.
    if selless:
        selless_block = (
            "<selless_order_data>\n"
            "REAL order data resolved from Selless (production) for this ticket. Treat as ground "
            "truth and fill the template from it (status, tracking, variant, amounts). Do NOT ask "
            "the customer for their order number — you already have the order.\n"
            f"{json.dumps(selless, ensure_ascii=False)}\n"
            "</selless_order_data>\n\n"
        )
    else:
        selless_block = (
            "<selless_order_data>\n"
            "No Selless order could be resolved for this ticket (empty/unknown order code). Per D-34, "
            "treat this as a signal: draft a verify-order / clarify-order-info reply that politely asks "
            "for the order number or checkout email. Never fabricate order facts.\n"
            "</selless_order_data>\n\n"
        )
    # Deterministic sub-type -> allowed template codes (honors the blocking
    # "free-pick template" anti-pattern: in headless 1-turn the model cannot run
    # subtype_to_code(), so the allowed set is injected and the code MUST come
    # from it for the classified sub-type — never invented, never cross-family).
    codes_map = "\n".join(
        f"  {st}: {', '.join(codes) if codes else '(NO template — use a clarify/none flow)'}"
        for st, (codes, _files) in sorted(_SUBTYPE_TEMPLATES.items())
    )
    return (
        "Process this customer support ticket through the full cs-agent-team pipeline "
        "(classify -> extract -> ground -> draft -> self-critique). The pipeline is "
        "ALWAYS-DRAFT (D-33): you ALWAYS return action=\"draft\" with a ready-to-send reply. "
        "There is NO escalate verdict. Ground the reply ONLY in the local file-store template "
        "you select (by Customer_Request sub-type) + the Selless order data below (D-29/D-31); never "
        "fabricate order facts (D-34).\n\n"
        "HARD CONSTRAINTS (CS-agent handling is the gold standard — match it):\n"
        "1) TEMPLATE FAMILY — after you set customer_request, template_code MUST be exactly one "
        "code from the ALLOWED list for that sub-type below (verbatim, e.g. \"B2\", \"G4\"). "
        "NEVER invent a code and NEVER use a code from another family (a G-code is shipping/"
        "delivery; an F-code is cancellation; do not mix). If the sub-type has no template, say so.\n"
        "2) SELLESS-FIRST BRANCH — read po_status and the delivery status from the Selless data "
        "FIRST and let it pick the flow/code: NEW/PROCESSING (not yet shipped) -> G2; "
        "TO/TA/in-transit -> G4 (neutral) or G5 (angry); DELIVERED -> delivered/in-transit update "
        "(or G10 if the customer says not received). Do not default to G2.\n"
        "3) FIT/SIZE COMPLAINT on a delivered item within guarantee -> customer_request = "
        "\"Replace\" (offer free replacement, keep items, ask for measurements if missing) — NOT "
        "Return (money-back only when the customer explicitly demands it) and NOT "
        "Change_Product_Variant (pre-delivery only).\n"
        "4) VERIFY CLAIM vs ORDER — if the customer claims a wrong item/variant, compare against "
        "the Selless ordered variant before classifying; if the order shows they received what they "
        "ordered, do not treat it as a fulfillment error.\n\n"
        "ALLOWED TEMPLATE CODES BY SUB-TYPE:\n" + codes_map + "\n\n"
        "Return EXACTLY ONE JSON object with BOTH your "
        "classification properties AND your final draft verdict:\n\n"
        "{\n"
        '  "properties": {\n'
        '    "category": "complaint|change_request|inquiry|other",\n'
        '    "customer_request": "<level-2 Customer_Request sub-type, e.g. Return/Replace/'
        'Partial_Refund/Full_Refund/Review/Cancel_Order/Change_Shipping_Address/'
        'Change_Product_Variant/Ask_About_Order/Ask_About_Delivery_Status/Ask_About_Policy/'
        'Ask_About_Product/Ask_About_Promotion>",\n'
        '    "confidence": "high|med|low",\n'
        '    "order_ref": "<extracted order ref or empty>",\n'
        '    "issue_type": "<short issue type>",\n'
        '    "product_line": "<if determinable, else empty>",\n'
        '    "template_code": "<the file-store template code you used, verbatim, e.g. B7/G2/F23; '
        'empty only if the sub-type has no template>",\n'
        '    "flow": "<workflow/flow name if determinable, else empty>",\n'
        '    "step": "<workflow step if determinable, else empty>",\n'
        '    "rootcause": "<root cause if determinable, else empty>",\n'
        '    "resolution_status": "<resolution status if determinable, else empty>",\n'
        '    "high_risk": true|false\n'
        "  },\n"
        '  "verdict": {"action":"draft","body":"<full customer reply text>",'
        '"template_code":"<same code as properties.template_code>",'
        '"escalation_hint": null  '
        '/* or {"reason":"money|legal|injection|low_confidence|missing_key","signals":{}} '
        'as ADVISORY only — it NEVER suppresses the draft */}\n'
        "}\n\n"
        f"ticket_id: {ticket.get('ticket_id', 'unknown')}\n"
        f"<ticket_metadata>\nsubject: {subject}\norder_ref: {order_ref}\n</ticket_metadata>\n\n"
        f"{selless_block}"
        f"<ticket_body>\n{body}\n</ticket_body>\n"
    )


def _iter_json_objects(text: str):
    """Yield every top-level balanced {...} block in *text*, parsed as JSON.

    Robust to markdown ```json fences and surrounding prose: the headless model
    often returns a narrative plus a fenced/embedded JSON object rather than a
    bare object. Brace-matching (string/escape aware) finds each candidate.
    """
    depth = 0
    start = -1
    in_str = False
    esc = False
    for i, ch in enumerate(text):
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
            continue
        if ch == '"':
            in_str = True
        elif ch == "{":
            if depth == 0:
                start = i
            depth += 1
        elif ch == "}":
            if depth > 0:
                depth -= 1
                if depth == 0 and start >= 0:
                    blob = text[start:i + 1]
                    try:
                        yield json.loads(blob)
                    except json.JSONDecodeError:
                        pass
                    start = -1


def _parse_combined(raw_output: str) -> dict:
    """Return {properties: dict, verdict: dict}.

    Always-draft (D-33): on a parse miss the verdict falls soft to
    action=draft with a parse_error hint (handled by the caller / _parse_verdict).
    The model's reply (claude --print --output-format json) wraps the model text
    in outer["result"]; that text may contain prose + a fenced JSON object, so we
    scan all balanced objects and pick the ones carrying our contract keys.
    """
    properties: dict = {}
    verdict: dict | None = None

    # 1. Unwrap the claude --print JSON envelope to get the model's text.
    inner_text = raw_output
    try:
        outer = json.loads(raw_output.strip())
        if isinstance(outer, dict) and isinstance(outer.get("result"), str):
            inner_text = outer["result"]
        elif isinstance(outer, dict) and "result" in outer:
            inner_text = json.dumps(outer["result"])
    except json.JSONDecodeError:
        pass

    # 2. Fast path: the whole inner text is a bare contract object.
    for candidate in (inner_text.strip(),):
        try:
            obj = json.loads(candidate)
        except json.JSONDecodeError:
            continue
        if isinstance(obj, dict) and ("properties" in obj or "action" in obj or "verdict" in obj):
            if isinstance(obj.get("properties"), dict):
                properties = obj["properties"]
            if isinstance(obj.get("verdict"), dict):
                verdict = obj["verdict"]
            elif "action" in obj:
                verdict = obj
            if properties or verdict:
                return {"properties": properties, "verdict": verdict or _parse_verdict(raw_output)}

    # 3. Robust path: scan every embedded JSON object (handles fences + prose).
    for obj in _iter_json_objects(inner_text):
        if not isinstance(obj, dict):
            continue
        if not properties and isinstance(obj.get("properties"), dict):
            properties = obj["properties"]
        if verdict is None and isinstance(obj.get("verdict"), dict):
            verdict = obj["verdict"]
        # A combined object that itself is the verdict (has action/body).
        if verdict is None and obj.get("action") == "draft":
            verdict = obj
        # A standalone properties object (has customer_request/category).
        if not properties and ("customer_request" in obj or "category" in obj) and "action" not in obj:
            properties = obj

    if verdict is None:
        verdict = _parse_verdict(raw_output)  # always-draft fail-soft
    return {"properties": properties, "verdict": verdict}


def _build_classify_prompt(ticket: dict, selless: dict | None) -> str:
    """Pass-1: classify only. Selless data is included so the sub-type respects
    the order record (e.g. wrong-variant claims verified against ordered variant)."""
    body = redact_text(ticket.get("body", ""))
    subject = redact_text(ticket.get("subject", ""))
    sb = (f"<selless_order_data>\n{json.dumps(selless, ensure_ascii=False)}\n</selless_order_data>\n\n"
          if selless else "<selless_order_data>\n(no order resolved)\n</selless_order_data>\n\n")
    return (
        "Classify this customer support ticket. Return EXACTLY ONE JSON object, no prose:\n"
        '{"category":"complaint|change_request|inquiry|other",'
        '"customer_request":"<level-2 sub-type: Return/Replace/Partial_Refund/Full_Refund/Review/'
        'Cancel_Order/Change_Shipping_Address/Change_Non_Shipping_Address/Change_Product_Variant/'
        'Ask_About_Order/Ask_About_Delivery_Status/Ask_About_Policy/Ask_About_Product/'
        'Ask_About_Promotion>","order_ref":"<order code or empty>","high_risk":true|false}\n\n'
        "Rules (CS gold standard): a fit/size complaint on a DELIVERED item within guarantee = "
        "Replace (not Return, not Change_Product_Variant). Return = customer explicitly wants money "
        "back. Change_Product_Variant = pre-delivery variant swap. If the customer claims a wrong "
        "variant but the Selless order shows they received what they ordered, it is "
        "Change_Product_Variant (they picked wrong), not Replace. po_status=CANCELLED with an OOS/"
        "angry customer = Full_Refund, not Ask_About_Order.\n\n"
        f"<ticket_metadata>\nsubject: {subject}\n</ticket_metadata>\n\n"
        f"{sb}"
        f"<ticket_body>\n{body}\n</ticket_body>\n"
    )


def _build_draft_prompt2(ticket: dict, selless: dict | None, templates_text: str,
                         allowed_codes: list[str], subtype: str) -> str:
    """Pass-2: draft with the REAL template bodies injected. The sub-type is already
    decided; the model selects template_code from the allowed set and FILLS the reply
    from the actual template content (all mandatory elements: offers/discounts/refund
    confirmations)."""
    body = redact_text(ticket.get("body", ""))
    subject = redact_text(ticket.get("subject", ""))
    order_ref = redact_text(ticket.get("order_ref", ""))
    sb = (f"<selless_order_data>\nREAL order data (ground truth — fill status/tracking/variant/"
          f"amounts from it; never fabricate):\n{json.dumps(selless, ensure_ascii=False)}\n"
          f"</selless_order_data>\n\n" if selless else
          "<selless_order_data>\nNo order resolved — per D-34 draft a clarify-order reply asking "
          "for order number/checkout email; never fabricate.\n</selless_order_data>\n\n")
    allowed_str = ", ".join(allowed_codes) if allowed_codes else "(none — this sub-type has no template; draft a brief clarification)"
    return (
        f"You are drafting the customer reply for a ticket already classified as "
        f"customer_request = \"{subtype}\". ALWAYS-DRAFT (D-33): return action=\"draft\".\n\n"
        "Use the ACTUAL template library below. Steps:\n"
        f"1) Pick template_code = exactly ONE code from the ALLOWED set: {allowed_str}. "
        "It MUST be the code whose template best matches this ticket + the Selless status. "
        "Never invent a code, never use a code outside this set.\n"
        "2) FILL that template into a complete, ready-to-send reply — include EVERY mandatory "
        "element the template specifies (e.g. the offer, the % discount, refund-window "
        "confirmation, measurement request). Ground all order facts in the Selless data.\n"
        "3) Match how the CS team resolves it (decision + offer), not just the topic.\n\n"
        "<templates>\n" + templates_text + "\n</templates>\n\n"
        "Return EXACTLY ONE JSON object, no prose:\n"
        '{"properties":{"category":"...","customer_request":"' + subtype + '","confidence":"high|med|low",'
        '"order_ref":"...","issue_type":"...","product_line":"...","template_code":"<one allowed code>",'
        '"flow":"...","step":"...","rootcause":"...","resolution_status":"...","high_risk":true|false},'
        '"verdict":{"action":"draft","body":"<full filled reply>","template_code":"<same code>",'
        '"escalation_hint":null}}\n\n'
        f"ticket_id: {ticket.get('ticket_id','unknown')}\n"
        f"<ticket_metadata>\nsubject: {subject}\norder_ref: {order_ref}\n</ticket_metadata>\n\n"
        f"{sb}"
        f"<ticket_body>\n{body}\n</ticket_body>\n"
    )


async def _run_cli(prompt: str) -> tuple[int, str, str]:
    proc = await asyncio.create_subprocess_exec(
        *_CLAUDE_CLI, stdin=asyncio.subprocess.PIPE,
        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE, cwd=str(_REPO_ROOT))
    out_b, err_b = await proc.communicate(input=prompt.encode())
    return proc.returncode, out_b.decode(errors="replace"), err_b.decode(errors="replace")


async def run_ai_team(ticket: dict, selless: dict | None = None,
                      category: str | None = None) -> dict:
    """Run the real cs-agent-team on *ticket*; return {properties, verdict}.

    D-33 always-draft: verdict is always action=draft. On cli_error or injection,
    an advisory escalation_hint is attached — the draft is still returned.
    D-32: CS_RUN_ID, _state_file_path, and _post_screen_draft are GONE (deleted guard hooks).
    D-39: DRY_RUN only — never posts to Freshdesk.
    """
    assert settings.dry_run, "FATAL: settings.dry_run is False — aborting (no live posts allowed)."

    # D-14 advisory pre-screen (mirrors run_ticket — advisory under D-30, never blocks draft)
    is_inj, reason = _pre_screen_ticket(ticket)
    injection_hint = (
        {"reason": reason, "signals": {"injection": True}} if is_inj else None
    )

    def _cli_error(stage: str, err: str) -> dict:
        return {"properties": {}, "verdict": {
            "action": "draft", "body": "", "citations": [],
            "escalation_hint": {"reason": f"cli_error:{stage}",
                                "signals": {"stderr": redact_text(err[:200])}}}}

    try:
        # PASS 1 — classify (Selless-aware so the sub-type respects the order record).
        rc1, out1, err1 = await _run_cli(_build_classify_prompt(ticket, selless))
        if rc1 != 0:
            return _cli_error("classify", err1)
        cls = _parse_combined(out1)["properties"] or _parse_combined(out1)["verdict"] or {}
        subtype = (cls.get("customer_request") or "").strip()
        cat = category or ""
        if not cat:  # infer category from the classified macro-category
            cat = {"complaint": "complaint", "change_request": "change_request",
                   "inquiry": "inquiry"}.get((cls.get("category") or "").strip(), "inquiry")

        # PASS 2 — draft with the REAL template bodies for that sub-type injected.
        templates = _load_templates_for_subtype(subtype, cat) if subtype else _load_templates(cat)
        allowed = _allowed_codes_for_subtype(subtype)
        rc2, out2, err2 = await _run_cli(
            _build_draft_prompt2(ticket, selless, templates, allowed, subtype or "Ask_About_Order"))
        if rc2 != 0:
            return _cli_error("draft", err2)
        parsed = _parse_combined(out2)
        # Backfill classification fields from pass-1 if pass-2 omitted them.
        for k in ("category", "customer_request", "order_ref", "high_risk"):
            if not parsed["properties"].get(k) and cls.get(k) not in (None, ""):
                parsed["properties"][k] = cls.get(k)
        # Advisory injection hint (D-14) — never changes the draft.
        if injection_hint and parsed["verdict"].get("escalation_hint") is None:
            parsed["verdict"]["escalation_hint"] = injection_hint
        return parsed
    except Exception as exc:  # noqa: BLE001
        return {"properties": {}, "verdict": {
            "action": "draft", "body": "", "citations": [],
            "escalation_hint": {"reason": f"run_error:{exc}", "signals": {}}}}


# ---------------------------------------------------------------------------
# Collect
# ---------------------------------------------------------------------------

def _select(category: str, per_cat: int, only_tid: str | None) -> list[dict]:
    rows = list(csv.DictReader(open(_CSV[category], newline="", encoding="utf-8")))
    if only_tid:
        rows = [r for r in rows if (r.get("Ticket ID") or "").strip() == only_tid]
    else:
        rows = rows[:per_cat]
    return rows


async def collect(per_cat: int, only_tid: str | None, only_cat: str | None) -> None:
    env = _load_env_prd()
    domain, key = env["FRESHDESK_DOMAIN"], env["FRESHDESK_API_KEY"]
    cats = [only_cat] if only_cat else list(_CSV.keys())
    records: list[dict] = []
    sclient = httpx.AsyncClient(base_url=_SELLESS_BASE, timeout=20)  # D-34 Selless grounding
    with httpx.Client() as client:
        for cat in cats:
            rows = _select(cat, per_cat, only_tid)
            print(f"== {cat}: {len(rows)} ticket(s) ==", flush=True)
            for i, row in enumerate(rows, 1):
                tid = (row.get("Ticket ID") or "").strip()
                if not tid:
                    continue
                # Shared per-ticket pipeline (D-34 Selless grounding + DRY_RUN ai team).
                # collect() and run() both call _process_row so the two paths cannot drift.
                print(f"[{cat} {i}/{len(rows)}]", end=" ", flush=True)
                rec = await _process_row(row, client, sclient, domain, key, cat)
                records.append(rec)
                time.sleep(0.3)
    await sclient.aclose()
    with open(_DATA_PATH, "w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"DONE {len(records)} records -> {_DATA_PATH}", flush=True)


# ---------------------------------------------------------------------------
# Draft mode — supply the real template files as context (no MCP/Voyage/DB).
# Produces an actual filled reply + runs the D-26 authorized-offer guard on it.
# ---------------------------------------------------------------------------

_SNAP = _TICKET_DIR.parent  # .../snapshots/confluence
_TEMPLATE_DIR = _REPO_ROOT / ".planning" / "phases" / \
    "01-knowledge-survey-conflict-inventory" / "snapshots"

_CATEGORY_TEMPLATE_GLOBS = {
    "complaint": ["product complaint-*.md", "shipping queries & complaints-*.md"],
    "change_request": ["cancellation request-*.md", "change request-*.md", "change-request-*.md"],
    "inquiry": ["shipping queries & complaints-*.md", "situational-*.md", "billing-*.md"],
}
_MAX_TEMPLATE_CHARS = 26000

# --- Snapshot files grouped by code family (verbatim names from snapshots/) ---
_F_A = "product complaint-within guarantee-template1.md"          # A1-A9
_F_B = "product complaint-within guarantee-template2.md"          # B1-B13
_F_D = "product complaint-within guarantee-template3.md"          # D1-D9 (follow-ups)
_F_C1 = "product complaint-out of guarantee-template.md"          # C1
_F_C2 = "product complaint-replacement not fit-template.md"       # C2
_F_E_CONTACT = "change request-template1.md"                      # E8 (email/phone), E9 (billing)
_F_E_VARIANT = "change-request-template2.md"                      # E4,E5,E10,E11,E12
_F_E_ADDR = "change request-template3.md"                         # E1,E13
_F_E_TA = "change request-template4.md"                           # E2,E6 (need SCE)
_F_E_TO = "change request-template5.md"                           # E3,E7 (cannot change)
_F_F_PROMO = "cancellation request-template1.md"                  # F23 (aftersale promotion)
_CANCEL_FILES = [f"cancellation request-template{n}.md" for n in range(2, 10)]  # F1-F22
_F_G_DNR = "shipping queries & complaints-template1.md"           # G10,G14,G15
_F_G_OOS = "shipping queries & complaints-template2.md"           # G3.1,G3.2
_F_G_RTS = "shipping queries & complaints-template3.md"           # G11,G13
_F_G_TEST = "shipping queries & complaints-template4.md"          # G12
_F_G_COMMON = "shipping queries & complaints-template5.md"        # G1,G2,G4-G9
_F_SITU = "situational-template.md"
_F_BILL = "billing-template.md"

# Deterministic Customer_Request sub-type -> (allowed template codes, snapshot files to load).
# Source of truth: CODE-MAP-templates.md. The drafter MUST pick template_used from `codes`
# (verbatim, e.g. "F23"); inventing a "Custom-..." name is rejected. `codes=[]` means the
# sub-type has NO dedicated template in the library (gap) -> draft a clarification, flag it.
_SUBTYPE_TEMPLATES: dict[str, tuple[list[str], list[str]]] = {
    # Inquiry / shipping
    "Ask_About_Delivery_Status": (
        ["G1", "G2", "G3.1", "G3.2", "G4", "G5", "G6", "G7", "G8", "G9", "G10", "G11", "G13", "G14", "G15"],
        [_F_G_COMMON, _F_G_DNR, _F_G_OOS, _F_G_RTS],
    ),
    "Ask_About_Order": (["G1", "G2", "G12"], [_F_G_COMMON, _F_G_TEST]),
    "Ask_About_Promotion": (["F23"], [_F_F_PROMO, _F_SITU]),
    "Ask_About_Policy": ([], [_F_SITU, _F_BILL]),
    "Ask_About_Product": ([], [_F_SITU]),
    # Complaint / product
    # B-codes (non-defective fit/sizing) listed FIRST: most Replace/Return tickets are fit
    # issues, not defects. A-codes only when the item is defective/wrong/missing/damaged.
    "Replace": (["B1", "B2", "B5", "B6", "A1", "A2", "A6", "A7", "A8", "A3", "C2"], [_F_B, _F_A, _F_C2]),
    "Return": (["B5", "B6", "B3", "B7", "A5", "A6", "A8", "A4", "A9", "D3", "D7"], [_F_B, _F_A, _F_D]),
    "Full_Refund": (["A4", "A9", "B3", "B7", "B13", "D3"], [_F_A, _F_B, _F_D]),
    "Review": ([], [_F_SITU, _F_C1]),  # GAP: no dedicated Review template (CODE-MAP gap)
    # Change request
    "Cancel_Order": ([f"F{n}" for n in range(1, 23)], _CANCEL_FILES),
    "Change_Shipping_Address": (["E1", "E2", "E3", "E13"], [_F_E_ADDR, _F_E_TA, _F_E_TO]),
    "Change_Product_Variant": (["E4", "E5", "E6", "E7", "E10", "E11", "E12"], [_F_E_VARIANT, _F_E_TA, _F_E_TO]),
    "Change_Non_Shipping_Address": (["E8", "E9"], [_F_E_CONTACT]),
}


def _subtype_of(rec: dict) -> str:
    return (rec.get("cs_props", {}).get("Customer_Request") or "").strip()


def _allowed_codes_for_subtype(subtype: str) -> list[str]:
    entry = _SUBTYPE_TEMPLATES.get(subtype)
    return list(entry[0]) if entry else []


def _read_template_files(files: list[str]) -> str:
    """Concatenate the given snapshot files (de-duped, char-capped)."""
    parts: list[str] = []
    total = 0
    for name in dict.fromkeys(files):  # preserve order, drop dups
        fp = _TEMPLATE_DIR / name
        if not fp.exists():
            continue
        txt = fp.read_text(encoding="utf-8", errors="replace")
        chunk = f"\n===== TEMPLATE FILE: {fp.name} =====\n{txt}\n"
        if total + len(chunk) > _MAX_TEMPLATE_CHARS:
            chunk = chunk[: max(0, _MAX_TEMPLATE_CHARS - total)]
        parts.append(chunk)
        total += len(chunk)
        if total >= _MAX_TEMPLATE_CHARS:
            break
    return "".join(parts)


def _load_templates_for_subtype(subtype: str, category: str) -> str:
    """Deterministic: load ONLY the snapshot files mapped to this sub-type.

    Falls back to the (broad) category glob if the sub-type is unknown.
    """
    entry = _SUBTYPE_TEMPLATES.get(subtype)
    if entry:
        return _read_template_files(entry[1])
    return _load_templates(category)


def _load_templates(category: str) -> str:
    parts: list[str] = []
    total = 0
    for pat in _CATEGORY_TEMPLATE_GLOBS.get(category, []):
        for fp in sorted(_TEMPLATE_DIR.glob(pat)):
            txt = fp.read_text(encoding="utf-8", errors="replace")
            chunk = f"\n===== TEMPLATE FILE: {fp.name} =====\n{txt}\n"
            if total + len(chunk) > _MAX_TEMPLATE_CHARS:
                chunk = chunk[: max(0, _MAX_TEMPLATE_CHARS - total)]
            parts.append(chunk)
            total += len(chunk)
            if total >= _MAX_TEMPLATE_CHARS:
                return "".join(parts)
    return "".join(parts)


def _build_draft_prompt(rec: dict, templates: str, selless: dict | None,
                        allowed_codes: list[str] | None = None) -> str:
    p = rec["cs_props"]
    cs_ctx = {k: p.get(k, "") for k in (
        "Subject", "Customer_Request", "Order", "Product_line", "Feedback_Issue",
        "Rootcause", "Section_Flow", "Status", "Tags") if (p.get(k) or "").strip()}
    msg = redact_text(rec.get("customer_msg", ""))[:5000]
    selless_block = (json.dumps(selless, ensure_ascii=False)[:3000] if selless
                     else "(no live Selless order data available for this ticket)")
    subtype = _subtype_of(rec)
    if allowed_codes:
        code_rule = (
            f"- ALLOWED TEMPLATE CODES for sub-type '{subtype}': {', '.join(allowed_codes)}.\n"
            "- You MUST set \"template_used\" to EXACTLY ONE code from that ALLOWED list "
            "(verbatim code only, e.g. \"F23\"). Pick the one whose template heading best fits the "
            "customer's situation (warranty/evidence/measurements/order-status as applicable).\n"
            "- Do NOT invent a 'Custom-...' name and do NOT use a code outside the ALLOWED list. "
            "Then FILL that template's body.\n"
        )
    else:
        # Gap sub-type: no dedicated template in the library.
        code_rule = (
            f"- NOTE: sub-type '{subtype}' has NO dedicated template in the library. "
            "Set \"template_used\" to \"NONE-<sub-type>\" and write a correct, on-brand "
            "clarification/acknowledgement reply grounded in the provided context — do NOT invent "
            "a fake template code or an unauthorized offer.\n"
        )
    return (
        "You are the REPLY DRAFTER for a US e-commerce customer-support team (Shophelp / "
        "RosyLift shapewear & apparel). This is a PoC. Your job: write the BEST customer reply "
        "for the ticket below. You MUST ALWAYS return a complete, ready-to-send reply — never "
        "refuse, never escalate, never leave the reply empty.\n\n"
        "How to draft:\n"
        f"{code_rule}"
        "- Use the CS workflow properties + live Selless order data to fill order-specific fields. "
        "Where a real value is genuinely unknown, keep the template's [placeholder] token (e.g. "
        "[tracking_number], [order_id], [name]) rather than inventing a value.\n"
        "- Include ONLY the refund/discount/replacement/retention offer that the CHOSEN template "
        "specifies — do not add an offer the template does not contain.\n"
        "- Write the complete email body in the 'reply' field, in the brand's polite CS voice.\n\n"
        "GROUNDING DISCIPLINE (mandatory — these caused real errors, follow exactly):\n"
        "1. NEVER claim an operational action was already done — EVEN IF the template text says "
        "so. Do NOT write 'we've changed / cancelled / updated / refunded / processed your order' "
        "or state a specific refund dollar amount you cannot ground. Rephrase to a request, an "
        "OFFER, or a next step ('we can change…', 'would you like…', 'we'll process the refund "
        "once you confirm…'). The team CANNOT execute order mutations.\n"
        "2. When there is NO live Selless order data, do NOT assume the order's fulfillment state "
        "(do NOT assume it already shipped / is undeliverable / is out of warranty). Pick the "
        "FIRST-RESPONSE / information-gathering code in the allowed set (e.g. confirm details, ask "
        "for measurements, share status + ETA) — NOT the highest-remediation / 'cannot change' / "
        "'shipped' branch — and offer only the modest retention/goodwill that branch authorizes. "
        "Do NOT volunteer a 100%/50% refund or a big discount the situation does not call for.\n"
        "3. If no order can be confirmed (no order number, order not found, or the customer's "
        "product/store clearly is NOT this brand), do NOT fabricate an order status. Draft a "
        "polite VERIFY-ORDER / CLARIFY reply asking for the order number or noting a possible "
        "wrong-store mix-up (still pick the closest allowed code, or NONE-<sub-type> if none fits).\n"
        "4. For a Replace/Return complaint about FIT / SIZING / SATISFACTION (item not defective), "
        "use a B-code (B1 ask measurements, B2 sizing-advice replacement) — NOT an A-code and NOT "
        "a 100% refund on first contact. Use A-codes ONLY when the item is genuinely DEFECTIVE / "
        "wrong / missing / damaged. First response = offer a replacement / ask for measurements; "
        "reserve refunds for later steps or explicit policy, and never exceed what the chosen "
        "code authorizes.\n\n"
        "Return EXACTLY ONE JSON object (no prose, no markdown fence):\n"
        "{\n"
        '  "classification": {"category":"complaint|change_request|inquiry|other",'
        '"customer_request":"<sub-type>","confidence":"high|med|low"},\n'
        '  "template_used": "<one ALLOWED code, verbatim>",\n'
        '  "offer": {"refund_pct":<n>,"discount_pct":<n>,"replacement":true|false,'
        '"retention_pct":<n>} OR null (only the dimensions the chosen template actually offers),\n'
        '  "reply": "<full ready-to-send email reply text>"\n'
        "}\n\n"
        f"TICKET ticket_id={rec['ticket_id']} category={rec['category_file']} "
        f"sub_type={subtype}\n"
        f"CS workflow properties: {json.dumps(cs_ctx, ensure_ascii=False)}\n"
        f"Live Selless order data: {selless_block}\n"
        f"<customer_message>\n{msg}\n</customer_message>\n\n"
        f"TEMPLATE LIBRARY (only the templates valid for sub-type '{subtype}'):\n{templates}\n"
    )


def _parse_draft_json(raw_output: str) -> dict | None:
    """Unwrap claude --output-format json envelope, then extract the model's JSON object."""
    try:
        outer = json.loads(raw_output.strip())
        inner = outer.get("result") if isinstance(outer, dict) else None
    except json.JSONDecodeError:
        inner = raw_output
    if isinstance(inner, dict):
        return inner
    if not isinstance(inner, str):
        return None
    s = inner.strip()
    # strip markdown fences if present
    if s.startswith("```"):
        s = s.split("```", 2)[1] if "```" in s[3:] else s
        s = s.lstrip("json").strip("` \n")
    # find the largest balanced {...}
    start = s.find("{")
    if start == -1:
        return None
    for end in range(len(s), start, -1):
        frag = s[start:end]
        if frag.rstrip().endswith("}"):
            try:
                return json.loads(frag)
            except json.JSONDecodeError:
                continue
    return None


def _summarize_selless(po: dict, dos: list) -> dict:
    """Compact, drafter-friendly grounding summary from /po/{id}.

    Surfaces the fields the template-branch decision needs: order status, DO
    fulfillment status (TA/TO/delivered…) + status dates, tracking numbers, and
    the ordered variant(s). Avoids dumping the full raw payload.
    """
    out = {
        "order_code": po.get("code"),
        "po_status": po.get("status"),
        "created": po.get("created"),
        "amount": po.get("amount"),
        "discount": po.get("discount"),
        "shipping_city": (po.get("shipping_address") or {}).get("city"),
        "shipping_state": (po.get("shipping_address") or {}).get("state"),
        "deliveries": [],
    }
    for d in (dos or [])[:5]:
        v = d.get("variant") or {}
        props = {p.get("name"): p.get("value") for p in (v.get("properties") or [])}
        out["deliveries"].append({
            "do_status": d.get("status"),
            "odo_status": d.get("odo_status"),
            "product": v.get("title") or d.get("product_label"),
            "variant": props,
            "trackings": d.get("trackings") or [],
            "date_processing": d.get("status_date_processing"),
            "date_ta": d.get("status_date_ta"),
            "date_delivered": d.get("status_date_delivered"),
            "date_cancelled": d.get("status_date_cancelled"),
        })
    return out


async def fetch_selless_order(client: httpx.AsyncClient, order_code: str) -> dict | None:
    """Resolve the human order code (e.g. "25659-2952") to live order detail.

    Path: GET /po/search?param=<code> -> internal id -> GET /po/{id}. Returns a
    compact summary, or None when the order genuinely is not found (empty search
    result = a real 'no order' signal the drafter should treat as verify/clarify).
    """
    order_code = (order_code or "").strip()
    if len(order_code) < 3:
        return None
    try:
        rs = await client.get("/po/search", params={"param": order_code, "skip": 0, "take": 1})
        if rs.status_code != 200:
            return None
        data = rs.json()
        items = data if isinstance(data, list) else (data.get("items") or [])
        if not items:
            return None  # genuine no-order: order code not in Selless
        oid = items[0].get("id")
        if not oid:
            return None
        ro = await client.get(f"/po/{oid}")
        if ro.status_code != 200:
            return None
        body = ro.json() or {}
        return _summarize_selless(body.get("po") or {}, body.get("dos") or [])
    except Exception:  # noqa: BLE001
        return None


async def run_drafter(rec: dict, templates: str, selless: dict | None = None,
                      allowed_codes: list[str] | None = None) -> dict:
    """Call claude --print headless with templates-as-context; return the drafted reply dict."""
    prompt = _build_draft_prompt(rec, templates, selless, allowed_codes)
    proc = await asyncio.create_subprocess_exec(
        *_CLAUDE_CLI,
        stdin=asyncio.subprocess.PIPE,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
        cwd=str(_REPO_ROOT),
    )
    out_b, err_b = await proc.communicate(input=prompt.encode())
    if proc.returncode != 0:
        return {"_error": f"cli_rc={proc.returncode}: {err_b.decode(errors='replace')[:200]}"}
    parsed = _parse_draft_json(out_b.decode(errors="replace"))
    return parsed if parsed is not None else {"_error": "parse_error"}


_SELLESS_BASE = settings.selless_api_base_url  # single source of truth (src/config.py: selless_env)


async def draft(only_cat: str | None) -> None:
    """Debug draft mode (D-35 deprecated for validation — use collect() for fidelity).

    Always-draft (D-33): produces a reply grounded on local templates + workflow
    (CS properties) + best-effort live Selless data. No guard, no escalation.
    This path is a debug/shortcut aid only; collect() via the real team is the
    D-35 validation path.
    """
    recs = [json.loads(l) for l in _DATA_PATH.read_text(encoding="utf-8").splitlines() if l.strip()]
    tmpl_cache: dict[str, str] = {}
    n_draft = n_err = n_selless = n_bad_code = 0
    async with httpx.AsyncClient(base_url=_SELLESS_BASE, timeout=20) as sclient:
        for i, rec in enumerate(recs, 1):
            cat = rec["category_file"]
            if only_cat and cat != only_cat:
                continue
            subtype = _subtype_of(rec)
            allowed = _allowed_codes_for_subtype(subtype)
            # cache key = sub-type (deterministic load); fall back to category
            ck = subtype or cat
            templates = tmpl_cache.setdefault(ck, _load_templates_for_subtype(subtype, cat))
            selless = await fetch_selless_order(sclient, rec.get("cs_props", {}).get("Order", ""))
            if selless:
                n_selless += 1
            d = await run_drafter(rec, templates, selless, allowed)
            if "_error" in d:
                rec["ai_draft"] = {"error": d["_error"], "reply": ""}
                n_err += 1
                print(f"[{i}/{len(recs)}] {cat} {rec['ticket_id']} -> ERROR {d['_error']}", flush=True)
                continue
            cls = d.get("classification", {}) or {}
            # Validate the chosen template code against the allowed set for this sub-type.
            tmpl_used = (d.get("template_used") or "").strip()
            code = tmpl_used.split("-", 1)[0].strip()  # "F23-Can cancel-..." -> "F23"
            if allowed:
                code_ok = code in allowed
            else:
                code_ok = tmpl_used.upper().startswith("NONE")  # gap sub-type: expect NONE-*
            if not code_ok:
                n_bad_code += 1
            rec["ai_properties"] = {
                "category": cls.get("category", ""),
                "customer_request": cls.get("customer_request", ""),
                "confidence": cls.get("confidence", ""),
            }
            rec["ai_draft"] = {
                "template_used": tmpl_used,
                "template_code": code,
                "template_valid": code_ok,
                "allowed_codes": allowed,
                "offer": d.get("offer"),
                "selless_used": bool(selless),
                "reply": d.get("reply", ""),
            }
            n_draft += 1
            off = d.get("offer")
            print(f"[{i}/{len(recs)}] {cat} {rec['ticket_id']} -> DRAFT "
                  f"cr={cls.get('customer_request','?')} tmpl={tmpl_used or '?'} "
                  f"{'OK' if code_ok else 'BAD-CODE!'} "
                  f"offer={'yes' if off else 'none'} selless={'Y' if selless else 'n'} "
                  f"reply_len={len(d.get('reply',''))}", flush=True)
            time.sleep(0.2)
    with open(_DATA_PATH, "w", encoding="utf-8") as f:
        for r in recs:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"DONE draft: {n_draft} drafted, {n_err} error, {n_bad_code} bad-template-code, "
          f"{n_selless} with live Selless data -> {_DATA_PATH}", flush=True)


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

def build_xlsx() -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    records = [json.loads(line) for line in _DATA_PATH.read_text(encoding="utf-8").splitlines() if line.strip()]
    wb = Workbook()
    wb.remove(wb.active)

    hdr_font = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    ai_fill = PatternFill("solid", fgColor="FCE4D6")
    wrap_top = Alignment(wrap_text=True, vertical="top")
    label_font = Font(bold=True, color="1F4E78")

    for rec in records:
        tid = rec["ticket_id"]
        ws = wb.create_sheet(title=tid[:31])
        props = rec["cs_props"]
        ai = rec.get("ai_properties", {}) or {}
        verdict = rec.get("ai_verdict", {}) or {}
        dft = rec.get("ai_draft", {}) or {}

        # Row 1 — Properties table header
        ws["A1"], ws["B1"], ws["C1"], ws["D1"] = (
            "Property", "CS Agent value", "AI Team value", "Conversation")
        for c in ("A1", "B1", "C1", "D1"):
            ws[c].font = hdr_font
            ws[c].fill = hdr_fill

        # AI value counterparts for the CS property rows (D-37 side-by-side compare)
        ai_for_cs = {
            "Customer_Request": ai.get("customer_request", ""),
            "Order": ai.get("order_ref", ""),
            "Product_line": ai.get("product_line", ""),
            "Flow": ai.get("flow", ""),
            "STEP": ai.get("step", ""),
            "Section_Flow": ai.get("flow", ""),
            "Rootcause": ai.get("rootcause", ""),
            "Rootcause_type": ai.get("rootcause", ""),
            "Resolution status": ai.get("resolution_status", ""),
        }

        # Build the ordered, non-empty CS property rows
        ordered = [p for p in _PRIORITY_PROPS if (props.get(p) or "").strip()]
        rest = [k for k, v in props.items()
                if (v or "").strip() and k not in ordered and k != "Ticket ID"]
        rows = ordered + rest

        r = 2
        for name in rows:
            ws.cell(r, 1, name)
            ws.cell(r, 2, props.get(name, ""))
            if name in ai_for_cs and str(ai_for_cs[name]).strip():
                ws.cell(r, 3, ai_for_cs[name])
            r += 1

        # AI-only properties block
        r += 1
        ws.cell(r, 1, "— AI Team output —").font = label_font
        r += 1
        offer = dft.get("offer") if dft else None
        if dft:
            tv = dft.get("template_valid")
            tv_str = "✓ valid" if tv else ("✗ INVALID (not in allowed set)" if tv is False else "")
            ai_rows = [
                ("AI: category", ai.get("category", "")),
                ("AI: customer_request", ai.get("customer_request", "")),
                ("AI: confidence", ai.get("confidence", "")),
                ("AI: template_used", dft.get("template_used", "")),
                ("AI: template_valid", tv_str),
                ("AI: allowed_codes", ", ".join(dft.get("allowed_codes") or []) or "(none — gap)"),
                ("AI: offer", json.dumps(offer, ensure_ascii=False) if offer else "(none)"),
                ("AI: live Selless data", "yes" if dft.get("selless_used") else "no"),
            ]
        else:
            hint = verdict.get("escalation_hint")
            ai_rows = [
                ("AI: category", ai.get("category", "")),
                ("AI: customer_request", ai.get("customer_request", "")),
                ("AI: confidence", ai.get("confidence", "")),
                ("AI: template_code", ai.get("template_code", "") or verdict.get("template_code", "")),
                ("AI: flow", ai.get("flow", "")),
                ("AI: step", ai.get("step", "")),
                ("AI: rootcause", ai.get("rootcause", "")),
                ("AI: resolution_status", ai.get("resolution_status", "")),
                ("AI: order_ref", ai.get("order_ref", "")),
                ("AI: issue_type", ai.get("issue_type", "")),
                ("AI: high_risk", str(ai.get("high_risk", ""))),
                ("AI: Selless grounded", "yes" if rec.get("selless_order") else "no"),
                ("AI: verdict_action", verdict.get("action", "")),
                ("AI: escalation_hint",
                 json.dumps(hint, ensure_ascii=False) if hint else "(none)"),
            ]
        for name, val in ai_rows:
            ws.cell(r, 1, name).font = label_font
            ws.cell(r, 3, val)
            ws.cell(r, 3).fill = ai_fill
            r += 1

        # Checker block — why AI differs from CS (from the diff-checker agent)
        chk = rec.get("checker") or {}
        if chk:
            r += 1
            ws.cell(r, 1, "— Checker (AI vs CS) —").font = label_font
            r += 1
            sev = (chk.get("severity") or "").lower()
            sev_fill = PatternFill("solid", fgColor={
                "error": "FFC7CE", "minor": "FFEB9C", "ok": "C6EFCE"}.get(sev, "FFFFFF"))
            chk_rows = [
                ("Checker: severity", chk.get("severity", "")),
                ("Checker: template_ok", chk.get("template_appropriate", "")),
                ("Checker: offer_grounded", chk.get("offer_grounded", "")),
                ("Checker: asserts_action", chk.get("asserts_operational_action", "")),
                ("Checker: diff vs CS", chk.get("diff", "")),
                ("Checker: reason", chk.get("reason", "")),
            ]
            for name, val in chk_rows:
                ws.cell(r, 1, name).font = label_font
                cell = ws.cell(r, 3, val)
                cell.alignment = wrap_top
                if name == "Checker: severity":
                    cell.fill = sev_fill
                r += 1

        # Column D — D2 customer msg, D3 CS reply, D4 AI reply (labels prefixed)
        if dft:
            ai_reply = f"[ERROR] {dft.get('error')}" if dft.get("error") else (dft.get("reply") or "[empty]")
        else:
            ai_reply = (verdict.get("body") if verdict.get("action") == "draft"
                        else f"[ESCALATE] reason={verdict.get('reason', '')} "
                             f"signals={json.dumps(verdict.get('signals', {}), ensure_ascii=False)}")
        fetch_err = rec.get("fetch_error", "")
        d2 = "▶ CUSTOMER — first message:\n\n" + (rec.get("customer_msg") or (f"[fetch error: {fetch_err}]" if fetch_err else "[empty]"))
        d3 = "▶ CS AGENT — first reply (actual):\n\n" + (rec.get("cs_reply") or "[no public reply found]")
        d4 = "▶ AI TEAM — drafted reply (PoC):\n\n" + (ai_reply or "[empty]")
        for cell, text in (("D2", d2), ("D3", d3), ("D4", d4)):
            ws[cell] = text
            ws[cell].alignment = wrap_top

        # Sizing
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 90
        for rr in (2, 3, 4):
            ws.row_dimensions[rr].height = 150

    wb.save(_XLSX_PATH)
    print(f"WROTE {len(records)} sheet(s) -> {_XLSX_PATH}", flush=True)


# ---------------------------------------------------------------------------
# Run subcommand helpers (D-41..D-45)
# ---------------------------------------------------------------------------

def _parse_ticket_list(path: str) -> list[dict]:
    """Parse a uat_ticket.csv (semicolon-delimited) or a plain one-ID-per-line file.

    For the semicolon-delimited format (header: Level_in;Resolved date;Ticket ID):
      - Preserves 'Ticket ID' and 'Level_in' bucket for each row.
      - 'Resolved date' is informational and kept but not used by the run path.

    For a plain one-ID-per-line file (no header row matching the expected columns):
      - Each non-empty line is treated as a Ticket ID with Level_in = "unknown".

    Returns list of dicts with at minimum: {'Ticket ID': str, 'Level_in': str}.
    """
    text = Path(path).read_text(encoding="utf-8", errors="replace")
    lines = [l.rstrip("\r\n") for l in text.splitlines() if l.strip()]
    if not lines:
        return []

    # Detect CSV format by checking if first line is the semicolon-header
    first = lines[0]
    if ";" in first and "Ticket ID" in first and "Level_in" in first:
        # Semicolon-delimited format
        reader = csv.DictReader(iter(lines), delimiter=";")
        rows: list[dict] = []
        for row in reader:
            tid = (row.get("Ticket ID") or "").strip()
            if not tid:
                continue
            rows.append({
                "Ticket ID": tid,
                "Level_in": (row.get("Level_in") or "").strip(),
                "Resolved date": (row.get("Resolved date") or "").strip(),
            })
        return rows

    # Plain one-ID-per-line (convenience format)
    rows = []
    for line in lines:
        tid = line.strip()
        if tid:
            rows.append({"Ticket ID": tid, "Level_in": "unknown"})
    return rows


def _apply_caps(
    rows: list[dict],
    limit: int | None,
    per_cat: int | None,
) -> tuple[list[dict], dict[str, int]]:
    """Apply --limit and --per-cat caps to a list of ticket rows.

    Args:
        rows: list of dicts with at minimum 'Level_in' and 'Ticket ID'.
        limit: total cap across all buckets (None = no total cap).
        per_cat: per-Level_in cap (None = no per-bucket cap).

    Returns:
        (selected_rows, dropped_report) where:
          - selected_rows: the rows that passed both caps.
          - dropped_report: dict mapping bucket -> dropped count.
    """
    # Apply per-cat cap first (bucket-level)
    if per_cat is not None:
        by_bucket: dict[str, list[dict]] = {}
        for r in rows:
            b = r.get("Level_in", "unknown") or "unknown"
            by_bucket.setdefault(b, []).append(r)

        selected: list[dict] = []
        dropped_report: dict[str, int] = {}
        for bucket, bucket_rows in by_bucket.items():
            kept = bucket_rows[:per_cat]
            dropped = bucket_rows[per_cat:]
            selected.extend(kept)
            if dropped:
                dropped_report[bucket] = len(dropped)
    else:
        selected = list(rows)
        dropped_report = {}

    # Apply global limit cap
    if limit is not None and len(selected) > limit:
        over = selected[limit:]
        selected = selected[:limit]
        # Attribute dropped to their buckets
        for r in over:
            b = r.get("Level_in", "unknown") or "unknown"
            dropped_report[b] = dropped_report.get(b, 0) + 1

    return selected, dropped_report


async def _process_row(
    row: dict,
    client: httpx.Client,
    sclient: httpx.AsyncClient,
    domain: str,
    key: str,
    category_hint: str | None = None,
) -> dict:
    """Per-ticket pipeline shared by both `collect()` and `run()`.

    Fetches conversation (Freshdesk GET, read-only), resolves Selless order (read-only
    when an order code is present), runs the real ai team (DRY_RUN), and returns a
    record dict in the same shape collect() accumulates.
    """
    tid = (row.get("Ticket ID") or "").strip()
    order_code = (row.get("Order") or "").strip()
    conv = fetch_conversation(client, domain, key, tid)
    ticket = {
        "ticket_id": tid,
        "subject": row.get("Subject", ""),
        "order_ref": order_code,
        "body": conv["customer_msg"],
    }
    # Resolve Selless order only when we have a code (uat_ticket.csv has no Order column
    # so order_code will be empty there -> None triggers D-34 clarify-order flow).
    selless = await fetch_selless_order(sclient, order_code) if order_code else None

    # category_hint: Level_in from uat_ticket.csv when present, else None
    ai = await run_ai_team(ticket, selless, category_hint or None)

    cat_label = category_hint or "unknown"
    act = ai["verdict"].get("action")
    cr = ai["properties"].get("customer_request", "?")
    tc = ai["properties"].get("template_code") or ai["verdict"].get("template_code", "?")
    print(
        f"  tid={tid} cat={cat_label} fetch={'ok' if not conv['error'] else conv['error']} "
        f"msg_len={len(conv['customer_msg'])} selless={'Y' if selless else 'n'} "
        f"AI={act} cr={cr} tmpl={tc}",
        flush=True,
    )
    return {
        "category_file": cat_label,
        "ticket_id": tid,
        "cs_props": {k: row.get(k, "") for k in row.keys()},
        "customer_msg": conv["customer_msg"],
        "cs_reply": conv["cs_reply"],
        "fetch_error": conv["error"],
        "selless_order": selless,
        "ai_properties": ai["properties"],
        "ai_verdict": ai["verdict"],
    }


async def run(
    ticket_id: str | None,
    list_path: str | None,
    limit: int | None,
    per_cat: int,
) -> None:
    """D-41: run subcommand — drive any ticket(s) through the real cs-agent-team.

    --id  <ticket_id>: single ticket by ID (one synthetic row, DRY_RUN, read-only PROD).
    --list <csv>: batch from uat_ticket.csv (;-delimited, Level_in bucket, D-42).
    --limit N: total cap (D-43).
    --per-cat N: per-Level_in cap, default 10 (D-43).

    Cap drops are always logged (count + buckets), never silent.
    Output: overwrites _DATA_PATH + calls build_xlsx() -> test-tickets.xlsx (D-44).
    Never POSTs to Freshdesk (assert settings.dry_run, D-39).
    """
    assert settings.dry_run, "FATAL: settings.dry_run is False — aborting (no live posts allowed, D-39)."

    env = _load_env_prd()
    domain, key = env["FRESHDESK_DOMAIN"], env["FRESHDESK_API_KEY"]

    # Build the row set
    if ticket_id:
        rows: list[dict] = [{"Ticket ID": ticket_id.strip(), "Level_in": ""}]
    elif list_path:
        all_rows = _parse_ticket_list(list_path)
        rows, dropped_report = _apply_caps(all_rows, limit=limit, per_cat=per_cat)
        if dropped_report:
            total_dropped = sum(dropped_report.values())
            bucket_summary = ", ".join(f"{b}={n}" for b, n in dropped_report.items())
            print(
                f"[run] Cap applied: {len(rows)} selected, {total_dropped} dropped "
                f"(per-bucket: {bucket_summary})",
                flush=True,
            )
        else:
            print(f"[run] {len(rows)} ticket(s) to process (no cap drops)", flush=True)
    else:
        print("ERROR: --id or --list is required for the run subcommand.", file=sys.stderr)
        sys.exit(1)

    records: list[dict] = []
    sclient = httpx.AsyncClient(base_url=_SELLESS_BASE, timeout=20)
    with httpx.Client() as client:
        for i, row in enumerate(rows, 1):
            tid = (row.get("Ticket ID") or "").strip()
            if not tid:
                continue
            # Use Level_in as a category hint when available (maps to collect() cat parameter)
            level_in = (row.get("Level_in") or "").strip()
            # Normalize Level_in to the collect() category keys (lower + underscore)
            cat_hint: str | None = None
            if level_in:
                normalized = level_in.lower().replace(" ", "_")
                # Map uat_ticket.csv Level_in values to collect() category keys
                cat_map = {
                    "change_request": "change_request",
                    "complaint": "complaint",
                    "inquiry": "inquiry",
                }
                cat_hint = cat_map.get(normalized)
                if cat_hint is None:
                    # Pass as-is for unknown categories (non-blocking)
                    cat_hint = normalized if normalized else None

            print(f"[run {i}/{len(rows)}] processing tid={tid} ...", flush=True)
            rec = await _process_row(row, client, sclient, domain, key, cat_hint)
            records.append(rec)
            time.sleep(0.3)

    await sclient.aclose()

    with open(_DATA_PATH, "w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"[run] DONE {len(records)} record(s) -> {_DATA_PATH}", flush=True)
    build_xlsx()


def main() -> int:
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    c = sub.add_parser("collect")
    c.add_argument("--per-cat", type=int, default=10)
    c.add_argument("--ticket", default=None)
    c.add_argument("--category", default=None, choices=list(_CSV.keys()))
    d = sub.add_parser("draft")
    d.add_argument("--category", default=None, choices=list(_CSV.keys()))
    sub.add_parser("xlsx")
    # D-41: new `run` subcommand
    r = sub.add_parser("run")
    r.add_argument("--id", dest="ticket_id", default=None, help="Single ticket ID")
    r.add_argument("--list", dest="list_path", default=None,
                   help="Path to uat_ticket.csv (;-delimited, header Level_in;Resolved date;Ticket ID)")
    r.add_argument("--limit", type=int, default=None, help="Total ticket cap (D-43)")
    r.add_argument("--per-cat", type=int, default=10,
                   help="Per-Level_in cap, default 10 (D-43)")
    args = ap.parse_args()
    if args.cmd == "collect":
        asyncio.run(collect(args.per_cat, args.ticket, args.category))
    elif args.cmd == "draft":
        asyncio.run(draft(args.category))
    elif args.cmd == "xlsx":
        build_xlsx()
    elif args.cmd == "run":
        asyncio.run(run(args.ticket_id, args.list_path, args.limit, args.per_cat))
    return 0


if __name__ == "__main__":
    sys.exit(main())
